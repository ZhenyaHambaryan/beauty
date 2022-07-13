import json
from notifications.views import send_email
from .serializers import AboutUsSerializer, Category2Serializer, CategorySerializer, CitySerializer, CountrySerializer, FeedbackSerializer, GeneralSerializer, ServiceSerializer, GenderSerializer, LanguageSerializer, TarifSerializer
from rest_framework import viewsets
from .models import AboutUs, Category, City, Country, Departement, Feedback, General, Service, Gender, Language, Tarif
from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import datetime
from django_filters import rest_framework as filters
from rest_framework.filters import SearchFilter, OrderingFilter 
from rest_framework.pagination import PageNumberPagination
import openpyxl, requests

class GenderViewSet(viewsets.ModelViewSet):
    queryset = Gender.objects.all()
    serializer_class = GenderSerializer

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

    @action(methods=['GET'], detail=False,url_path='categories-with-services', url_name='categories_with_services')
    def categories_with_services(self, request,pk=None):
        return Response(Category2Serializer(Category.objects.filter(
                                    services__isnull=False).distinct(),many=True).data,status=200)

class LanguageViewSet(viewsets.ModelViewSet):
    queryset = Language.objects.all()
    serializer_class = LanguageSerializer

class ServiceViewSet(viewsets.ModelViewSet):
    serializer_class = ServiceSerializer

    def get_queryset(self):
        keyword = self.request.query_params.get('search','')
        category = self.request.query_params.get('category','')
        queryset = Service.objects.filter(Q(name_en__icontains=keyword) | 
                                             Q(name_fr__icontains=keyword)).distinct()
        if category is not None and category != "":
            queryset = queryset.filter(category_id=category)
        return queryset

class TarifViewSet(viewsets.ModelViewSet):
    queryset = Tarif.objects.filter(is_deleted=False)
    serializer_class = TarifSerializer
    permission_classes =[IsAuthenticated]

    @action(methods=['GET'], detail=False,url_path='tarif-list', url_name='tarif_list')
    def tarif_list(self, request,pk=None):
        paginator = PageNumberPagination()
        paginator.page_size = request.GET.get("limit",10)
        qs = Tarif.objects.filter(is_active=True, is_deleted=False).order_by("month")
        
        tarifs = paginator.paginate_queryset(qs, request)
        tr = []
        hr = []
        for tarif in tarifs:
            subscription = tarif.tarif_users.filter(user__user_id=request.user.id, 
                                                    expire_date__gte = datetime.datetime.now(),
                                                    payed=True)
            if subscription.count()>0:
                tr.append({
                    "expiration_date":subscription.first().expire_date,
                    "subscribed":True,
                    "tarif":TarifSerializer(tarif).data
                })
            else:
                hr.append({
                    "expiration_date":None,
                    "subscribed":False,
                    "tarif":TarifSerializer(tarif).data
                })
            arr = sorted(tr, key = lambda i: i['expiration_date'],reverse=True)
        
        return paginator.get_paginated_response(arr+hr)

    @action(methods=['DELETE'], detail=True,url_path='delete-tarif', url_name='delete_tarif')
    def delete_tarif(self, request,pk=None):
        tarif = Tarif.objects.get(id=pk)
        if tarif.tarif_users.all().count()>0:
            tarif.is_deleted=True
            tarif.is_active=False
            tarif.save()
        else:
            tarif.delete()
        
        return Response({"message":"OK"},status=204)
            
        
        
class FeedbackViewSet(viewsets.ModelViewSet):
    queryset = Feedback.objects.all().order_by("is_seen","-id")
    serializer_class = FeedbackSerializer

    def create(self, request, *args, **kwargs):
        email = request.data['email']
        phone_number = request.data['phone_number']
        if Feedback.objects.filter(email=email, created_at__date=datetime.date.today()).count()>=3 or\
                    Feedback.objects.filter(phone_number=phone_number, created_at__date=datetime.date.today()).count()>=3 :
            return Response({"message":"Your limit is expired for today."},status=400)
        feedback = Feedback(email=email,
                            phone_number=phone_number,
                            message=request.data['message'],
                            name=request.data['name'])
        feedback.save()
        send_email(to_email=email, lang="FR", code="FDBM")
        return Response(FeedbackSerializer(feedback).data, status=201)

class AboutUsViewSet(viewsets.ModelViewSet):
    queryset = AboutUs.objects.all().order_by('order')
    serializer_class = AboutUsSerializer
    filter_backends = [filters.DjangoFilterBackend]
    filter_fields = ['is_active',]

class CountryViewSet(viewsets.ModelViewSet):
    queryset = Country.objects.all().order_by("title_en","title_fr")
    serializer_class = CountrySerializer

class CityViewSet(viewsets.ModelViewSet):
    queryset = City.objects.all().order_by("title_en","title_fr")
    serializer_class = CitySerializer
    filter_backends = [filters.DjangoFilterBackend, SearchFilter,]
    filter_fields = ['country',]
    search_fields = ['title_en','title_fr']

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_popular_services(request):
    all = []
    services = ServiceSerializer(Service.objects.filter(
                    is_popular=True).order_by("?")[:5],many=True).data
    all+=services

    if len(services)<5:
        all += ServiceSerializer(Service.objects.filter(
                        is_popular=False).order_by("?")[:5-len(services)],many=True).data
    return Response(all)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def activate_deactivate_tarif(request,pk):
    tarif = Tarif.objects.get(id=pk)
    tarif.is_active = not tarif.is_active
    tarif.save()
    return Response({"message":"OK"},status=200)

@api_view(['GET'])
def set_feedback_seen(request,pk):
    feedback = Feedback.objects.get(id=pk)
    feedback.is_seen = True
    feedback.save()
    return Response(FeedbackSerializer(feedback).data, status=201)


class GeneralViewSet(viewsets.ModelViewSet):
    queryset = General.objects.all()
    serializer_class = GeneralSerializer

    def list(self, request):
        queryset = General.objects.all().first()
        serializer = GeneralSerializer(queryset)
        return Response(serializer.data)

@api_view(['POST'])
def upload_departements(request):
    excel_file = request.FILES["excel_file"]
    country_id = request.data['country']
    wb = openpyxl.load_workbook(excel_file)
    sheetname = str(wb.sheetnames[0])
    try:
        worksheet = wb[sheetname]
    except:
        return Response({"message": "file format or excel sheet name is incorrect."}, status=400)
    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value).strip())
        excel_data.append(row_data)
    for index, row in enumerate(excel_data[1:]):
        if row[0] !="None": 
            item = Departement()
            item.title_en = row[2]
            item.title_fr = row[2]
            item.number = row[1]
            item.country_id = country_id
            item.save()

    return Response({'message': "OK"})

@api_view(['POST'])
def upload_cities(request):
    excel_file = request.FILES["excel_file"]
    country_id = request.data['country']
    wb = openpyxl.load_workbook(excel_file)
    sheetname = str(wb.sheetnames[0])
    try:
        worksheet = wb[sheetname]
    except:
        return Response({"message": "file format or excel sheet name is incorrect."}, status=400)
    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value).strip())
        excel_data.append(row_data)
    departements = Departement.objects.all()
    for index, row in enumerate(excel_data[1:]):
        if row[0] !="None": 
            item = City()
            item.title_en = row[5]
            item.title_fr = row[5]
            item.departement_id = departements.get(number = row[1]).id
            item.country_id = country_id
            item.logitude = row[9]
            item.latitude = row[10]
            item.save()

    return Response({'message': "OK"})

def send_sms(body, phonenumber):
    r = requests.post("https://sms.capitolemobile.com/api/sendsms/xml", 
                  data={"XML":"""<SMS>
                            <authentification>
                                <username>beautycils</username>
                            <password>a7fcfdb0f85429dac688083aedda25a485d23ea0</password>
                            </authentification>
                            <message>
                                <text>"""+body+"""</text>
                        <sender>Beautycils</sender>
                        <route></route>
                        <long></long>
                        <prog></prog>
                        <id></id>
                            </message>
                            <recipients>
                                <gsm>"""+phonenumber+"""</gsm>
                            </recipients>
                        </SMS>"""})
    return str(r.text)

